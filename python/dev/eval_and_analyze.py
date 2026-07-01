#!/usr/bin/env python3
"""GPU 评估 + 失败分析 → 直接输出 Excel
用法: python eval_and_analyze.py [--mode all|2only|both] [--samples N]
"""
import os, sys, argparse, json, random, time, torch
os.environ["PYTORCH_ALLOC_CONF"] = "expandable_segments:True"
from collections import defaultdict, Counter
from transformers import AutoModelForCausalLM, AutoTokenizer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

parser = argparse.ArgumentParser()
parser.add_argument("--mode", choices=["all","2only","both"], default="both")
parser.add_argument("--samples", type=int, default=10000)
parser.add_argument("--token-range", type=str, default="",
                    help="Token range to sweep, e.g. 1-20.")
parser.add_argument("--nofix", action="store_true",
                    help="Skip pin_fix filtering (test raw baseline)")
args = parser.parse_args()

# ============================================================
MODEL_PATH = "d:/modelscope_models/Qwen/Qwen3___5-0___8B-Base"
CORPUS_PATH = "D:/分词注音工程/分读音词频统计/data/sentences_new.txt"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DICT_PATH = os.path.join(SCRIPT_DIR, "..", "配置文件", "pdsp.dict.yaml")
PIN_FIX_PATH = os.path.join(SCRIPT_DIR, "..", "配置文件", "pin_fix.txt")
OUTPUT = os.path.join(SCRIPT_DIR, "eval_results.xlsx")

DEVICE = "cuda"; BATCH_SIZE = 48; FIXED_SEQ = 64
MAX_CANDIDATES = 4
TARGET = args.samples
RUN_ALL = args.mode in ("all", "both")
RUN_2ONLY = args.mode in ("2only", "both")

# Token sweep range
if args.token_range:
    lo, hi = args.token_range.split("-")
    TOKEN_SWEEP = list(range(int(lo), int(hi)+1))
    N_TOKENS = TOKEN_SWEEP[0]  # for pre-encode (use max needed)
else:
    TOKEN_SWEEP = [4]
    N_TOKENS = 4

random.seed(42)

# ============================================================
# 1. 字典 → 两种模式
# ============================================================
print("[1/3] 字典 + 固顶...")
code_all = defaultdict(list)
code_2only = defaultdict(list)

with open(DICT_PATH, 'r', encoding='utf-8') as f:
    in_header = True
    for line in f:
        if in_header:
            if line.rstrip() == '...': in_header = False; continue
        parts = line.strip().split('\t')
        if len(parts) < 2: continue
        w, c = parts[0].strip(), parts[1].strip()
        if not w or not c or len(c) != 4: continue
        code_all[c].append(w)
        if len(w.encode('utf-8')) == 6:
            code_2only[c].append(w)

def build(code_words):
    hom, whc = {}, defaultdict(list)
    for code, words in code_words.items():
        u = list(dict.fromkeys(words))
        if len(u) >= 2: hom[code] = u
        for w in u: whc[w].append(code)
    return hom, set(whc.keys()), whc

hom_all, search_all, whc_all = build(code_all)
hom_2only, search_2only, whc_2only = build(code_2only)

# 加载固顶编码
pin_fix_codes = set()
with open(PIN_FIX_PATH, 'r', encoding='utf-8') as f:
    for line in f:
        line = line.strip()
        if line and not line.startswith('#') and '\t' in line:
            pin_fix_codes.add(line.split('\t', 1)[0])
fix_info = "跳过" if args.nofix else str(len(pin_fix_codes))
print(f"  全词:{len(hom_all)} 仅二字词:{len(hom_2only)} 固顶:{fix_info} 模式:{args.mode} 样本:{TARGET}")

# ============================================================
# 2. 采样
# ============================================================
print(f"[2/3] 采样 (跳过固顶编码)...")
import jieba; jieba.setLogLevel(20)
test_all, test_2only = [], []
seen_all, seen_2only = set(), set()
need_all = RUN_ALL and TARGET
need_2only = RUN_2ONLY and TARGET

with open(CORPUS_PATH, 'r', encoding='utf-8') as f:
    for line in f:
        if len(test_all) >= need_all and len(test_2only) >= need_2only: break
        line = line.strip()
        if not line: continue
        try: words = list(jieba.cut(line))
        except: continue
        for i, w in enumerate(words):
            if len(w) < 2: continue
            prev = ''.join(words[max(0, i-20):i])
            if len(prev) < 2: continue

            if need_all and len(test_all) < need_all and w in search_all:
                codes = whc_all[w]
                ok = codes[0] in hom_all and (args.nofix or codes[0] not in pin_fix_codes)
                if ok:
                    k = (prev[-30:], w, 'A')
                    if k not in seen_all:
                        seen_all.add(k)
                        test_all.append((prev, w, codes[0], hom_all[codes[0]][:9]))

            if need_2only and len(test_2only) < need_2only and w in search_2only:
                codes = whc_2only[w]
                ok = codes[0] in hom_2only and (args.nofix or codes[0] not in pin_fix_codes)
                if ok:
                    k = (prev[-30:], w, 'B')
                    if k not in seen_2only:
                        seen_2only.add(k)
                        test_2only.append((prev, w, codes[0], hom_2only[codes[0]][:9]))

def baseline(pairs):
    if not pairs: return 0
    return sum(1 for _, w, _, c in pairs if c[0] == w) / len(pairs)

print(f"  全词:{len(test_all)} 基线{baseline(test_all):.1%}  仅二字词:{len(test_2only)} 基线{baseline(test_2only):.1%}")

# ============================================================
# 3. 模型 + 评估 + 记录失败
# ============================================================
print("[3/3] 评估...")
tokenizer = AutoTokenizer.from_pretrained(MODEL_PATH, local_files_only=True, trust_remote_code=True)
if tokenizer.pad_token_id is None: tokenizer.pad_token_id = tokenizer.eos_token_id
model = AutoModelForCausalLM.from_pretrained(MODEL_PATH, torch_dtype=torch.bfloat16, device_map="auto",
                                              local_files_only=True, trust_remote_code=True)
model.eval()
torch.set_float32_matmul_precision('high')

def pre_encode(pairs):
    ctx_list, cands_list = [], []
    for prev_text, _, _, candidates in pairs:
        ctx_ids = tokenizer.encode(prev_text, add_special_tokens=False)
        if len(ctx_ids) > 80: ctx_ids = ctx_ids[-80:]
        ctx_list.append(ctx_ids)
        cands_list.append([tokenizer.encode(w, add_special_tokens=False) or [tokenizer.eos_token_id]
                           for w in candidates[:MAX_CANDIDATES]])
    return ctx_list, cands_list

ctx_all, cands_all = pre_encode(test_all)
ctx_2, cands_2 = pre_encode(test_2only)

code_net = defaultdict(lambda: {"save": 0, "hurt": 0})  # save=LLM挽救, hurt=LLM误判

@torch.no_grad()
def evaluate_full(ctx_list, cands_list, pairs, label, n_tokens=None):
    """评估并记录每个样本的详细信息"""
    if n_tokens is None: n_tokens = N_TOKENS
    rows_data, offsets = [], [0]
    for ctx_ids, cand_ids in zip(ctx_list, cands_list):
        truncated = ctx_ids[-n_tokens:] if len(ctx_ids) >= n_tokens else ctx_ids
        ctx_len = len(truncated)
        for w_ids in cand_ids:
            rows_data.append((truncated, ctx_len, w_ids))
        offsets.append(len(rows_data))

    total_rows = len(rows_data)
    max_len = max(min(len(ctx) + len(w), FIXED_SEQ) for ctx, _, w in rows_data)
    input_ids = torch.full((total_rows, max_len), tokenizer.pad_token_id, dtype=torch.long, device=DEVICE)
    ctx_lens = torch.zeros(total_rows, dtype=torch.long, device=DEVICE)
    w_lens = torch.zeros(total_rows, dtype=torch.long, device=DEVICE)

    for i, (ctx_ids, ctx_len, w_ids) in enumerate(rows_data):
        total = min(ctx_len + len(w_ids), FIXED_SEQ)
        ctx_slot = min(ctx_len, total)
        input_ids[i, :ctx_slot] = torch.tensor(ctx_ids[:ctx_slot], device=DEVICE)
        cand_slot = min(len(w_ids), total - ctx_slot)
        if cand_slot > 0:
            input_ids[i, ctx_slot:ctx_slot + cand_slot] = torch.tensor(w_ids[:cand_slot], device=DEVICE)
        ctx_lens[i] = ctx_len; w_lens[i] = cand_slot

    all_scores = []
    for start in range(0, total_rows, BATCH_SIZE):
        end = min(start + BATCH_SIZE, total_rows)
        out = model(input_ids=input_ids[start:end], use_cache=False)
        logits = out.logits.float()
        for b in range(end - start):
            cl = ctx_lens[start + b].item(); wl = w_lens[start + b].item()
            if wl == 0: all_scores.append(-1e10); continue
            logits_c = logits[b, cl - 1:cl - 1 + wl]
            labels_c = input_ids[start + b, cl:cl + wl]
            ce = torch.nn.functional.cross_entropy(logits_c, labels_c, reduction='sum').item()
            all_scores.append(-ce)

    correct = 0
    failures = []
    for i in range(len(pairs)):
        start = offsets[i]; end = offsets[i + 1]
        if start >= len(all_scores): break
        scores = all_scores[start:end]
        if not scores: continue
        ranked = sorted(enumerate(scores), key=lambda x: -x[1])
        prev_text, correct_word, code, orig_cands = pairs[i]
        cands = orig_cands[:MAX_CANDIDATES]
        llm_pick = cands[ranked[0][0]] if ranked else "?"
        is_correct = (llm_pick == correct_word)
        dict_correct = (cands[0] == correct_word)
        if is_correct: correct += 1
        else:
            failure = {
                'code': code, 'correct': correct_word, 'llm_pick': llm_pick,
                'margin': round(ranked[0][1] - ranked[1][1], 2) if len(ranked) >= 2 else 0,
                'dict_ok': dict_correct,
                'context': prev_text[-60:],
                'correct_n_tok': len(tokenizer.encode(correct_word, add_special_tokens=False)),
                'llm_n_tok': len(tokenizer.encode(llm_pick, add_special_tokens=False)),
                'candidates': ' | '.join(cands[:4]),
                'ranking': ' → '.join(
                    f"{idx+1}.{cands[ci]}({sc:.1f}){'✓' if cands[ci]==correct_word else ''}"
                    for idx, (ci, sc) in enumerate(ranked)
                ),
            }
            failures.append(failure)

        # 统计每个编码的双向效果
        if is_correct and not dict_correct:
            code_net[code]["save"] += 1   # LLM 挽救
        elif not is_correct and dict_correct:
            code_net[code]["hurt"] += 1   # LLM 误判

    acc = correct / max(1, len(pairs))
    print(f"  {label}: {acc:.1%} 失败{len(failures)}  dict_ok_llm_fail={sum(1 for f in failures if f['dict_ok'])}")
    return acc, failures

# ── 跑选择的模式 ──
if len(TOKEN_SWEEP) > 1:
    @torch.no_grad()
    def evaluate_fast(ctx_list, cands_list, pairs, n_tokens):
        # 1. 批量收集所有序列 (纯 Python list，避免逐行 torch.tensor)
        seqs, ctx_lens_list, w_lens_list = [], [], []
        offsets = [0]
        for ctx_ids, cand_ids in zip(ctx_list, cands_list):
            truncated = ctx_ids[-n_tokens:] if len(ctx_ids) >= n_tokens else ctx_ids
            ctx_len = len(truncated)
            for w_ids in cand_ids:
                total = min(ctx_len + len(w_ids), FIXED_SEQ)
                ctx_slot = min(ctx_len, total)
                cand_slot = min(len(w_ids), total - ctx_slot)
                seq = truncated[:ctx_slot] + w_ids[:cand_slot]
                seqs.append(torch.tensor(seq, dtype=torch.long))
                ctx_lens_list.append(ctx_slot)
                w_lens_list.append(cand_slot)
            offsets.append(len(seqs))

        total_rows = len(seqs)
        # 2. 一次性 padding (C++ 实现，快)
        input_ids = torch.nn.utils.rnn.pad_sequence(seqs, batch_first=True,
                                                      padding_value=tokenizer.pad_token_id).to(DEVICE)
        ctx_lens = torch.tensor(ctx_lens_list, dtype=torch.long, device=DEVICE)
        w_lens = torch.tensor(w_lens_list, dtype=torch.long, device=DEVICE)

        # 3. GPU 推理
        all_scores = []
        for start in range(0, total_rows, BATCH_SIZE):
            end = min(start + BATCH_SIZE, total_rows)
            out = model(input_ids=input_ids[start:end], use_cache=False)
            logits = out.logits.float()
            for b in range(end - start):
                cl = ctx_lens[start + b].item(); wl = w_lens[start + b].item()
                if wl == 0: all_scores.append(-1e10); continue
                logits_c = logits[b, cl - 1:cl - 1 + wl]
                labels_c = input_ids[start + b, cl:cl + wl]
                ce = torch.nn.functional.cross_entropy(logits_c, labels_c, reduction='sum').item()
                all_scores.append(-ce)

        # 4. 统计 Hit@1
        correct = 0
        for i in range(len(pairs)):
            start = offsets[i]; end = offsets[i + 1]
            if start >= len(all_scores): break
            scores = all_scores[start:end]
            if not scores: continue
            best_idx = scores.index(max(scores))
            _, correct_word, _, orig_cands = pairs[i]
            cands = orig_cands[:MAX_CANDIDATES]
            if best_idx < len(cands) and cands[best_idx] == correct_word:
                correct += 1
        return correct / max(1, len(pairs))

    base_a = baseline(test_all); base_b = baseline(test_2only)
    print(f"\n  全词基线={base_a:.1%}  仅二字词基线={base_b:.1%}  样本={TARGET}")
    print(f"  {'tok':>4} {'全词':>8} {'Δ':>7} {'二字词':>8} {'Δ':>7} {'耗时':>8}")
    print(f"  {'-'*48}")
    all_results = {}
    t0 = time.perf_counter()
    for nt in TOKEN_SWEEP:
        t1 = time.perf_counter()
        row = f"  {nt:>4}"
        if RUN_ALL:
            acc = evaluate_fast(ctx_all, cands_all, test_all, nt)
            row += f" {acc:>7.1%} {acc-base_a:>+6.1%}"
            all_results.setdefault("全词", {})[nt] = acc
        if RUN_2ONLY:
            acc = evaluate_fast(ctx_2, cands_2, test_2only, nt)
            row += f" {acc:>7.1%} {acc-base_b:>+6.1%}"
            all_results.setdefault("仅二字词", {})[nt] = acc
        dt1 = time.perf_counter() - t1
        row += f" {dt1:>7.1f}s"
        print(row, flush=True)
    dt = time.perf_counter() - t0
    import json as _json
    _json.dump(all_results, open(os.path.join(SCRIPT_DIR, "eval_sweep.json"), "w"), indent=2)
    print(f"\n总耗时: {dt:.0f}s  结果已保存: eval_sweep.json", flush=True)
    sys.exit(0)

t0 = time.perf_counter()
modes = []
if RUN_ALL:
    t0_a = time.perf_counter()
    acc_a, fails_a = evaluate_full(ctx_all, cands_all, test_all, "全词")
    base_a = baseline(test_all)
    print(f"  全词            {base_a:>7.1%} {acc_a:>7.1%} {acc_a-base_a:>+7.1%} {len(test_all):>6} {len(fails_a):>6} {time.perf_counter()-t0_a:.0f}s")
    modes.append(("全词", test_all, fails_a, base_a, acc_a))
else:
    acc_a, fails_a = None, []
if RUN_2ONLY:
    t0_b = time.perf_counter()
    acc_b, fails_b = evaluate_full(ctx_2, cands_2, test_2only, "仅二字词")
    base_b = baseline(test_2only)
    print(f"  仅二字词         {base_b:>7.1%} {acc_b:>7.1%} {acc_b-base_b:>+7.1%} {len(test_2only):>6} {len(fails_b):>6} {time.perf_counter()-t0_b:.0f}s")
    modes.append(("仅二字词", test_2only, fails_b, base_b, acc_b))
else:
    acc_b, fails_b = None, []
print(f"  总耗时: {time.perf_counter()-t0:.0f}s")

# ============================================================
# 4. 生成 Excel
# ============================================================
print("\n生成 Excel...")
wb = Workbook()
HF = Font(bold=True, size=11, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="4472C4")
B = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
W = Alignment(wrap_text=True, vertical="top")
Y = PatternFill("solid", fgColor="FFF2CC")

def hdr(ws, cols, widths):
    for c, (h, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HF; cell.fill = HFILL; cell.border = B; cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

def rows_out(ws, rows, start=2):
    for r, row in enumerate(rows, start):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v); cell.border = B; cell.alignment = W

# ── Sheet 1: 总览 ──
ws1 = wb.active; ws1.title = "总览"
hdr(ws1, ["模式","样本","基线","LLM","提升","失败数","dict_ok→LLM错","都错"], [12,8,8,8,8,8,14,8])
for label, pairs, fails, base, acc in modes:
    d_ok = sum(1 for f in fails if f['dict_ok'])
    rows_out(ws1, [[label, len(pairs), f"{base:.1%}", f"{acc:.1%}", f"{acc-base:+.1%}",
                    len(fails), d_ok, len(fails)-d_ok]])

# ── Sheet 2/3: 失败详情 ──
fail_cols = ["编码","正确词","LLM选","分差","字典OK","正确tok","LLM tok","候选词","排名详情","上文"]
fail_widths = [8,12,12,8,8,8,8,30,50,60]

all_fails_for_pinfix = []
for label, pairs, fails, base, acc in modes:
    sheet_name = f"失败-{label}"
    all_fails_for_pinfix.extend(fails)
    ws = wb.create_sheet(sheet_name)
    hdr(ws, fail_cols, fail_widths)
    rows = []
    for f in sorted(fails, key=lambda x: -x['margin']):
        rows.append([f['code'], f['correct'], f['llm_pick'], f['margin'], f['dict_ok'],
                     f['correct_n_tok'], f['llm_n_tok'], f['candidates'], f['ranking'][:200], f['context']])
    rows_out(ws, rows)
    ws.auto_filter.ref = ws.dimensions

# ── Sheet 4: 固顶建议（净收益 = LLM误判 - LLM挽救，正数才值得固顶）──
ws4 = wb.create_sheet("固顶建议")
# 收集失败样例供上下文展示
code_ex = {}
for f in all_fails_for_pinfix:
    if f['dict_ok'] and f['code'] not in code_ex: code_ex[f['code']] = f

hdr(ws4, ["编码","净收益","LLM误判","LLM挽救","正确词(应固顶)","示例上文"], [8,8,10,10,16,60])
rows4 = []
for code, net in code_net.items():
    save = net["save"]; hurt = net["hurt"]
    net_benefit = hurt - save  # 正数=固顶有利，负数=LLM更好
    if hurt > 0:  # 只有被LLM误判过的编码才考虑
        ex = code_ex.get(code, {})
        rows4.append([code, net_benefit, hurt, save,
                      ex.get('correct','?'), ex.get('context','')])
rows4.sort(key=lambda r: -r[1])  # 按净收益降序
rows_out(ws4, rows4)
for r in range(2, len(rows4)+2):
    net = ws4.cell(row=r, column=2).value
    if net and net > 0:
        for c in range(1,7): ws4.cell(row=r, column=c).fill = Y  # 正收益高亮
ws4.auto_filter.ref = ws4.dimensions

try:
    wb.save(OUTPUT)
except PermissionError:
    alt = OUTPUT.replace(".xlsx", "_1.xlsx")
    wb.save(alt)
    OUTPUT = alt
print(f"已保存: {OUTPUT}")
